# Excersize
# Read xl file, inventory
# List each company with respective product count -> how many products after from each supplier
# List products with inventory less than 10
# List each company with respective total inventory value
# Calculate and write inventory value for each product into spreadsheet


# Built in module for all types of files

# external library special for spreadsheet -> openpyxl

# Module = single .py file
# Package = collection of modules must contain init.py file
# library = collection of packages

import openpyxl as xl

# Read xl file, inventory
inventory = xl.load_workbook('inventory.xlsx')
products = inventory['Sheet1']

# List each company with respective product count -> how many products from each supplier
print(products)
products_per_supplier = {}
print(products.max_row)  # number of rows in sheet #75
for product_row in range(2, products.max_row+1): # from 2 to less than 76
    supplier_name = products.cell(product_row, 4).value
    # print(supplier_name)
    if supplier_name not in products_per_supplier:
        print(f'adding new Supplier {supplier_name}')
        products_per_supplier[supplier_name] = 1
    else:
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name]+1

print(f'products from each supplier {products_per_supplier}')

# List products with inventory less than 10
products_with_inventory_less_than_10 = {}

for product_row in range(2, products.max_row+1):
    product_number = int(products.cell(product_row, 1).value)
    product_inventory = int(products.cell(product_row, 2).value)
    if product_number not in products_with_inventory_less_than_10 and product_inventory < 10:
        products_with_inventory_less_than_10[product_number] = product_inventory

print(f'List of products with inventory less than 10 {products_with_inventory_less_than_10}')

# List each company with respective total inventory value
company_inventory_value = {}
for product_row in range(2, products.max_row+1):
    inventory_price = products.cell(product_row, 2).value
    price = products.cell(product_row, 3).value
    supplier_name = products.cell(product_row, 4).value
    if supplier_name not in company_inventory_value:
        company_inventory_value[supplier_name] = price*inventory_price
    else:
        company_inventory_value[supplier_name] += price*inventory_price

print(f'each company with respective total inventory value {company_inventory_value}')

# Calculate and write inventory value for each product into spreadsheet
product_inventory_value = {}
for product_row in range(2, products.max_row+1):
    product_id = int(products.cell(product_row, 1).value)
    inventory_price = products.cell(product_row, 2).value
    price = products.cell(product_row, 3).value
    # updating the cell
    total_inventory_price = products.cell(product_row, 5)

    if product_id in product_inventory_value:
        product_inventory_value[product_id] += price*inventory_price
    else:
        product_inventory_value[product_id] = price*inventory_price
    # value is added in memory not in sheet
    total_inventory_price.value = price*inventory_price

# save the file
inventory.save('inventory_updated_with_inventory_price_col.xlsx')

print(f'each product with respective total inventory value {product_inventory_value}')